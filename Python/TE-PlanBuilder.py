import requests
import re
import sys
import readline
import json
import math
from datetime import datetime, timedelta
from time import *
from prettytable import PrettyTable
from enum import Enum
from threading import Timer
from openpyxl import load_workbook
import shutil

CMDs = ["show agents list enterprise", "show agents list cluster", "show agents list endpoint", "show agents list cloud", "show agents list non endpoint",
                             "show test list agent-server", "show test list http-server", "show test list all summary", "show test list dns-trace", "show test list dns-sec",
                             "show test list agent-server summary", "show test list http-server summary", "show test list dns-trace summary", "show test list dns-sec summary"]

token = ""
rate_limit = -1
rate_limit_threshold = 1
organization_max_rate = 250
account_group = ""
aid = 0
BASE_URL = "https://api.thousandeyes.com/v6"
TIMER_SECONDS = 60
headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + token,
}
session = requests.Session()

class EndpointLicense(Enum):
    ADVANTAGE = "advantage"
    ESSENTIALS = "essentials"
    EMBEDDED = "embedded"

class Watchdog(Exception):
    def __init__(self, timeout, userHandler=None):  # timeout in seconds
        self.timeout = timeout
        self.handler = userHandler if userHandler is not None else self.defaultHandler
        self.timer = Timer(self.timeout, self.handler)
        self.timer.start()

    def reset(self):
        self.timer.cancel()
        self.timer = Timer(self.timeout, self.handler)
        self.timer.start()

    def stop(self):
        self.timer.cancel()

    def defaultHandler(self):
        raise self

def ResetRequestCounter():
    global rate_limit
    if rate_limit <= organization_max_rate and rate_limit < rate_limit_threshold :
        rate_limit = -1
    watchdog.reset()
    sys.exit()

watchdog = Watchdog(TIMER_SECONDS,ResetRequestCounter)

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

class AgentType(Enum):
    CLOUD = 0,
    ENTERPRISE = 1,
    ENTERPRISE_CLUSTER = 2
    ENDPOINT = 3

class TestType(Enum):
    agent_to_server = "agent-to-server"
    agent_to_agent = "agent-to-agent"
    http_server = "http-server"
    page_load = "page-load"
    dns_trace = "dns-trace"
    dns_dnssec = "dns-dnssec"
    dns_server = "dns-server"
    voiceRTPStream = "voice"
    sip_server = "sip-server"
    scheduled_test = "endpoint-tests"
    automated_test = "endpoint-automated-session-tests"
    transaction_test = "web-transactions"

class DocumentType(Enum):
    ENTERPRISE_AGENTS = 1
    ENDPOINT_AGENTS = 2

def SelectColumnsFromPrettyTable(myTable:PrettyTable, columns):
    columns = [s.strip() for s in columns]
    fieldNames = [str(s).strip() for s in myTable.field_names]
    
    columns = [ x for x in columns if x in fieldNames]
    
    if len(columns) < 1:
        return myTable

    for tableColumn in fieldNames:
        if tableColumn not in columns:
            try:
                myTable.del_column(tableColumn)
            except:
                pass
    
    if myTable.sortby not in myTable.field_names:
        myTable.sortby = myTable.field_names[0]
    return myTable

def TestCredentials(myToken:str):
    global headers
    global rate_limit
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/usage.json"
    result = session.get(url= BASE_URL + get_resource, headers = headers)
    try:
        rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
    except Exception as ex:
        print(ex)
    code = result.status_code 

    if code > 199 and code < 400:
        return True
    return False

def ValidateAccountName(myToken:str, accountName:str, showCMD = False):
    global rate_limit
    global aid
    global organization_max_rate
    global account_group
    get_resource = "/account-groups.json"
    
    if rate_limit > rate_limit_threshold or rate_limit == -1:
        result = session.get(url= BASE_URL + get_resource, headers = headers)
        try:
            rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
            organization_max_rate = int(result.headers["x-organization-rate-limit-limit"])
        except Exception as ex:
            print(ex)
        if result.status_code > 199 and result.status_code < 300:
            if accountName != "":
                    for accountGroup in result.json()["accountGroups"]:
                        if accountGroup["accountGroupName"] == accountName:
                            aid = accountGroup["aid"]
                            account_group = accountName
                            return accountGroup["aid"]
            #If we got here, either we did not receive a name or the name does not exist
            listOfAccountGroups = []
            agTable = PrettyTable()
            agTable.field_names = ["Account Group Name", "Account ID", "Organization Name"]
            for accountGroup in result.json()["accountGroups"]:
                agTable.add_row([accountGroup["accountGroupName"], accountGroup["aid"], accountGroup["organizationName"]])
            
            dummyTable = PrettyTable()
            if showCMD == False:
                dummyTable.field_names = ["I did not receive a name or the name provided is not associated to the account"]
                dummyTable.add_row(["Please select a name, from the table below:"])
                print()
                print(dummyTable)
                agTable.sortby = "Account Group Name"
                print(agTable)
                accountName = input("Account Group Name: ")
                return ValidateAccountName(myToken, accountName)
            else:
                agTable.sortby = "Account Group Name"
                agTable.align = "l"
                return agTable
        else:
            return False
    else:
        return False

def GetListOfAgents(myToken:str, agentType:AgentType, aid:str, returnType="table", selectColumns="ALL"):
    global headers
    global rate_limit
    myTable = PrettyTable()
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/agents.json"
    results = []

    if rate_limit > rate_limit_threshold or rate_limit == -1:
        if agentType == AgentType.ENDPOINT:
            get_resource = "/endpoint-agents.json"
            myTable.field_names = ["Agent Name", "Public IP", "Private IPs", "Connection", "Computer Name", "Version", "Status", "Location", "Created Date"]
            result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)
        else:
            if agentType == AgentType.CLOUD:
                myTable.field_names = ["AgentID", "Agent Name", "Type", "Country", "Location"]
            elif agentType == AgentType.ENTERPRISE or agentType == AgentType.ENTERPRISE_CLUSTER:
                myTable.field_names = ["AgentID", "Agent Name", "Type", "Country", "Location", "Enabled", "Public IP", "Agent State", "Created Date"]
            result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid) + "&agentTypes=" + agentType.name, headers = headers)
        
        try:
            rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
        except Exception as ex:
            print(ex)
        
        keyword=""

        if "agents" in result.json():
            keyword = "agents"
        elif "endpointAgents" in result.json():
            keyword = "endpointAgents"

        for agent in result.json()[keyword]:
            enabled = ""
            ipAddresses = ""
            agentState = ""

            if agent["agentType"] == "Enterprise" or agent["agentType"] == "Enterprise Cluster":
                if('enabled' in agent):
                    if(agent['enabled'] == 1):
                        enabled = "Yes"
                    elif(agent['enabled'] == 0):
                        enabled = "No"
                if('publicIpAddresses' in agent):
                    ipAddresses = agent['publicIpAddresses']
                if('agentState' in agent):
                    agentState = agent['agentState']
                
                if("clusterMembers" in agent):
                    #Iterate through cluster members
                    for agentMember in agent["clusterMembers"]:
                        results.append(["Enterprise Agent", agentMember["name"], agent["agentName"], "", str(agent["countryId"]) + ", " + str(agent["location"]), agentMember["agentState"]])
                else:
                    results.append(["Enterprise Agent", agent["agentName"], "", "", str(agent["countryId"]) + ", " + str(agent["location"]), agent["agentState"]])
                
            elif agent["agentType"] == "Cloud":
                myTable.add_row([agent["agentId"],agent["agentName"],agent["agentType"],agent["countryId"],
                                agent["location"] ])
            elif agent["agentType"] == "enterprise" or agent["enterprise-pulse"]: 
                if agent["agentType"] == "enterprise":
                    agentType = "Endpoint"
                elif agent["agentType"] == "enterprise-pulse":
                    agentType = "Endpoint-Pulse"
                
                if "location" in agent:
                    results.append([agentType, agent["agentName"], "", "", str(agent["location"]['locationName']), agent["status"]])
                else:
                    results.append([agentType, agent["agentName"], "", "", "", agent["status"]])
                
    else:
        return
    
    if "Created Date" in myTable.field_names:
        myTable.sortby = "Created Date"
    elif "Country" in myTable.field_names:
        myTable.sortby = "Country"
    myTable.align = "l"
    
    return results
    code = result.status_code 

def GetAgentsFromTest(myToken:str, aid:str, testID):
    global headers
    global rate_limit
    myTable = PrettyTable()
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/tests/"
    listOfAgents = []

    if rate_limit > rate_limit_threshold or rate_limit == -1:
        result = session.get(url= BASE_URL + get_resource + str(testID) + ".json?aid=" + str(aid), headers = headers)
        if result.status_code == 429:
            print(bcolors.WARNING + "Reached Rate limit" + bcolors.ENDC)
            while(result.status_code == 429):
                result = session.get(url= BASE_URL + get_resource + str(testID) + ".json?aid=" + aid, headers = headers)
        try:
            rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
        except Exception as ex:
            print(ex)
        if result.status_code != 200:
            return ""

        test = result.json()["test"]
        if "agents" in test[0]:
            for agent in test[0]["agents"]:
                listOfAgents.append(str(agent["agentName"]))
        else:
            return ""
    return listOfAgents

def GetAgentFromID(myToken:str, ais:str, agentID):
    global headers
    global rate_limit
    myTable = PrettyTable()
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/agents.json"
    listOfAgents = []

    if rate_limit > rate_limit_threshold or rate_limit == -1:
        result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)
        if result.status_code == 429:
            print(bcolors.WARNING + "Reached Rate limit" + bcolors.ENDC)
            while(result.status_code == 429):
                result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)
        try:
            print(bcolors.OKCYAN + "Performed Request, remaining: " + str(rate_limit)+ " requests" + bcolors.ENDC)
            rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
        except Exception as ex:
            print(ex)
        if result.status_code != 200:
            return ""

        agentList = result.json()["agents"]
        if len(agentList) > 0:
            for agent in agentList:
                if(agent["agentId"]==agentID):
                    return agent["agentName"]
        
        return ""

def GetAlertsFromTest(myToken:str, aid:str, testID):
    global headers
    global rate_limit
    myTable = PrettyTable()
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/tests/"
    listOfAlerts = []

    if rate_limit > rate_limit_threshold or rate_limit == -1:
        result = session.get(url= BASE_URL + get_resource + str(testID) + ".json?aid=" + aid, headers = headers)
        if result.status_code == 429:
            print(bcolors.WARNING + "Reached Rate limit" + bcolors.ENDC)
            while(result.status_code == 429):
                result = session.get(url= BASE_URL + get_resource + str(testID) + ".json?aid=" + aid, headers = headers)
        try:
            rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
        except Exception as ex:
            print(ex)
        test = result.json()["test"]
        if "alertRules" in test[0]:
            for alert in test[0]["alertRules"]:
                listOfAlerts.append(str(alert["ruleName"]))
    return listOfAlerts

def GetTests(myToken:str, testType:TestType, aid:str, returnType="table", selectColumns="ALL"):
    global headers
    global rate_limit
    myTable = PrettyTable()
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/tests.json"
    results = []

    if rate_limit > rate_limit_threshold or rate_limit == -1:
        myTable = PrettyTable()
        result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)
        
        try:
            print(bcolors.OKCYAN + "Performed Request, remaining: " + str(rate_limit)+ " requests" + bcolors.ENDC)
            rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
        except Exception as ex:
            print(ex)
        
        keyword=""
        listOfTests = []

        for test in result.json()["test"]:
            domain = ""
            recordType = ""
            alertsEnabled = ""
            target = ""
            port = ""
            networkMeasurements = ""
            bwMeasurements = ""
            mtuMeasurements = ""
            bgpMeasurements = ""
            pingPayload = ""
            redirects = ""
            createdDate = ""
            enabled = ""

            if "savedEvent" in test:
                if test["savedEvent"] == 0:
                    if test["type"] == testType.value and test["type"] == TestType.dns_server.value:
                            if(test["enabled"] == 1):
                                enabled = "Enabled"
                            else:
                                enabled = "Disabled"
                            if('alertsEnabled' in test):
                                if test['alertsEnabled'] == 0:
                                    alertsEnabled = "No"
                                elif test['alertsEnabled'] == 1:
                                    alertsEnabled = "Yes"
                            if('domain' in test):
                                domain = str(test['domain']).split(" ")[0]
                                recordType = str(test['domain']).split(" ")[1]
                            if "createdDate" in test:
                                createdDate = test["createdDate"]
                            
                            alertString = ""
                            agentString = ""
                            #alerts = GetAlertsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                            agents = GetAgentsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                            
                            #for alert in alerts:
                            #    alertString += alert + ","
                            #alertString = alertString[:-1]
                            for agent in agents:
                                agentString += agent + ";"
                            agentString = agentString[:-1]

                            if ["dns-server", test["testName"], str(domain) + " " + str(test["dnsQueryClass"]) + " " + str(recordType),
                                                agentString, test["interval"], "","", "Completed", enabled ] not in results:
                                results.append(["dns-server", test["testName"], str(domain) + " " + str(test["dnsQueryClass"]) + " " + str(recordType),
                                                agentString, test["interval"], "","", "Completed", enabled ])
                    elif test["type"] == testType.value and test["type"] == TestType.dns_trace.value:
                            if(test["enabled"] == 1):
                                enabled = "Enabled"
                            else:
                                enabled = "Disabled"
                            if('alertsEnabled' in test):
                                if test['alertsEnabled'] == 0:
                                    alertsEnabled = "No"
                                elif test['alertsEnabled'] == 1:
                                    alertsEnabled = "Yes"
                            if('domain' in test):
                                domain = str(test['domain']).split(" ")[0]
                                recordType = str(test['domain']).split(" ")[1]
                            if "createdDate" in test:
                                createdDate = test["createdDate"]
                            
                            alertString = ""
                            agentString = ""
                            #alerts = GetAlertsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                            agents = GetAgentsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                            
                            #for alert in alerts:
                            #    alertString += alert + ","
                            #alertString = alertString[:-1]
                            for agent in agents:
                                agentString += agent + ";"
                            agentString = agentString[:-1]

                            if ["dns-trace", test["testName"], str(domain) + " " + str(recordType),
                                                agentString, test["interval"], "","", "Completed", enabled ] not in results:
                                results.append(["dns-trace", test["testName"], str(domain) + " " + str(recordType),
                                                agentString, test["interval"], "","", "Completed", enabled ])
                    elif test["type"] == testType.value and test["type"] == TestType.dns_dnssec.value:
                            if(test["enabled"] == 1):
                                enabled = "Enabled"
                            else:
                                enabled = "Disabled"
                            if('domain' in test):
                                domain = str(test['domain']).split(" ")[0]
                                recordType = str(test['domain']).split(" ")[1]
                            if test['alertsEnabled'] == 0:
                                alertsEnabled = "No"
                            elif test['alertsEnabled'] == 1:
                                alertsEnabled = "Yes"
                            if "createdDate" in test:
                                createdDate = test["createdDate"]
                            
                            alertString = ""
                            agentString = ""
                            #alerts = GetAlertsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                            agents = GetAgentsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                            
                            #for alert in alerts:
                            #    alertString += alert + ","
                            #alertString = alertString[:-1]
                            for agent in agents:
                                agentString += agent + ";"
                            agentString = agentString[:-1]

                            if ["dns-dnssec", test["testName"], str(domain) + " " + str(recordType),
                                                agentString, test["interval"], "","", "Completed", enabled ] not in results:
                                results.append(["dns-dnssec", test["testName"], str(domain) + " " + str(recordType),
                                                agentString, test["interval"], "","", "Completed", enabled ])
                    
                    elif test["type"] == testType.value and test["type"] == TestType.agent_to_server.value:
                            if(test["enabled"] == 1):
                                enabled = "Enabled"
                            else:
                                enabled = "Disabled"
                            if('server' in test):
                                if ":" in test['server']:
                                    target = str(test['server']).split(":")[0]
                                    port = str(test['server']).split(":")[1]
                                else:
                                    target = str(test['server']).split(":")[0]
                                    port = ""
                            if test['alertsEnabled'] == 0:
                                alertsEnabled = "No"
                            elif test['alertsEnabled'] == 1:
                                alertsEnabled = "Yes"
                            if test['networkMeasurements'] == 0:
                                networkMeasurements = "No"
                            elif test['networkMeasurements'] == 1:
                                networkMeasurements = "Yes"
                            if test['mtuMeasurements'] == 0:
                                mtuMeasurements = "No"
                            elif test['mtuMeasurements'] == 1:
                                mtuMeasurements = "Yes"
                            if test['bandwidthMeasurements'] == 0:
                                bwMeasurements = "No"
                            elif test['bandwidthMeasurements'] == 1:
                                bwMeasurements = "Yes"
                            if test['bgpMeasurements'] == 0:
                                bgpMeasurements = "No"
                            elif test['bgpMeasurements'] == 1:
                                bgpMeasurements = "Yes"
                            if test['pingPayloadSize'] == -1:
                                pingPayload = "Auto"
                            else:
                                pingPayload = test['pingPayloadSize']
                            if "createdDate" in test:
                                createdDate = test["createdDate"]
                            
                            alertString = ""
                            agentString = ""
                            #alerts = GetAlertsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                            agents = GetAgentsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                            
                            #for alert in alerts:
                            #    alertString += alert + ","
                            #alertString = alertString[:-1]
                            for agent in agents:
                                agentString += agent + ";"
                            agentString = agentString[:-1]

                            if ["agent-to-server", test["testName"], target,
                                                agentString, test["interval"], "","", "Completed", enabled ] not in results:
                                results.append(["agent-to-server", test["testName"], target,
                                                agentString, test["interval"], "","", "Completed", enabled ])
                    elif test["type"] == testType.value and test["type"] == TestType.agent_to_agent.value:
                            if(test["enabled"] == 1):
                                enabled = "Enabled"
                            else:
                                enabled = "Disabled"
                            if "createdDate" in test:
                                createdDate = test["createdDate"]
                            
                            alertString = ""
                            agentString = ""
                            #alerts = GetAlertsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                            agents = GetAgentsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                            target = GetAgentFromID(myToken, aid, test["targetAgentId"])
                            
                            #for alert in alerts:
                            #    alertString += alert + ","
                            #alertString = alertString[:-1]
                            for agent in agents:
                                agentString += agent + ";"
                            agentString = agentString[:-1]

                            if ["agent-to-agent", test["testName"], target,
                                                agentString, test["interval"], "","", "Completed", enabled ] not in results:
                                results.append(["agent-to-agent", test["testName"], target,
                                                agentString, test["interval"], "","", "Completed", enabled ])
                    elif test["type"] == testType.value and test["type"] == TestType.http_server.value:
                            if(test["enabled"] == 1):
                                enabled = "Enabled"
                            else:
                                enabled = "Disabled"
                            if test['alertsEnabled'] == 0:
                                alertsEnabled = "No"
                            elif test['alertsEnabled'] == 1:
                                alertsEnabled = "Yes"
                            if test['networkMeasurements'] == 0:
                                networkMeasurements = "No"
                            elif test['networkMeasurements'] == 1:
                                networkMeasurements = "Yes"
                            if test['mtuMeasurements'] == 0:
                                mtuMeasurements = "No"
                            elif test['mtuMeasurements'] == 1:
                                mtuMeasurements = "Yes"
                            if test['bandwidthMeasurements'] == 0:
                                bwMeasurements = "No"
                            elif test['bandwidthMeasurements'] == 1:
                                bwMeasurements = "Yes"
                            if test['bgpMeasurements'] == 0:
                                bgpMeasurements = "No"
                            elif test['bgpMeasurements'] == 1:
                                bgpMeasurements = "Yes"
                            if test['followRedirects'] == 0:
                                redirects = "No"
                            elif test['followRedirects'] == 1:
                                redirects = "Yes"
                            if "createdDate" in test:
                                createdDate = test["createdDate"]
                            
                            alertString = ""
                            agentString = ""
                            #alerts = GetAlertsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                            agents = GetAgentsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                            
                            #for alert in alerts:
                            #    alertString += alert + ","
                            #alertString = alertString[:-1]
                            for agent in agents:
                                agentString += agent + ","
                            agentString = agentString[:-1]
                            if ["http-server", test["testName"], test["url"],
                                agentString, test["interval"], "","", "Completed", enabled ] not in results:
                                results.append(["http-server", test["testName"], test["url"],
                                                agentString, test["interval"], "","", "Completed", enabled ])
                    elif test["type"] == testType.value and test["type"] == TestType.page_load.value:
                            if(test["enabled"] == 1):
                                enabled = "Enabled"
                            else:
                                enabled = "Disabled"
                            if test['alertsEnabled'] == 0:
                                alertsEnabled = "No"
                            elif test['alertsEnabled'] == 1:
                                alertsEnabled = "Yes"
                            if test['networkMeasurements'] == 0:
                                networkMeasurements = "No"
                            elif test['networkMeasurements'] == 1:
                                networkMeasurements = "Yes"
                            if test['mtuMeasurements'] == 0:
                                mtuMeasurements = "No"
                            elif test['mtuMeasurements'] == 1:
                                mtuMeasurements = "Yes"
                            if test['bandwidthMeasurements'] == 0:
                                bwMeasurements = "No"
                            elif test['bandwidthMeasurements'] == 1:
                                bwMeasurements = "Yes"
                            if test['bgpMeasurements'] == 0:
                                bgpMeasurements = "No"
                            elif test['bgpMeasurements'] == 1:
                                bgpMeasurements = "Yes"
                            if test['followRedirects'] == 0:
                                redirects = "No"
                            elif test['followRedirects'] == 1:
                                redirects = "Yes"
                            if "createdDate" in test:
                                createdDate = test["createdDate"]
                            
                            alertString = ""
                            agentString = ""
                            #alerts = GetAlertsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                            agents = GetAgentsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                            
                            #for alert in alerts:
                            #    alertString += alert + ","
                            #alertString = alertString[:-1]
                            for agent in agents:
                                agentString += agent + ","
                            agentString = agentString[:-1]
                            if ["page-load", test["testName"], test["url"],
                                agentString, test["interval"], "","", "Completed", enabled ] not in results:
                                results.append(["page-load", test["testName"], test["url"],
                                                agentString, test["interval"], "","", "Completed", enabled ])
                                
                            if [test["testName"], testType.value, test["url"], test['interval'] , alertsEnabled, test['pageLoadTimeLimit'], test['pageLoadTargetTime'],
                                            test["httpTimeLimit"], test["httpTargetTime"], networkMeasurements, bwMeasurements, bgpMeasurements, mtuMeasurements,
                                            test["protocol"], test["probeMode"], test['pathTraceMode'], test["numPathTraces"], test['sslVersion'],
                                            test["verifyCertificate"], test['authType'], test['httpVersion'], redirects, createdDate, test["enabled"], alertString, agentString] not in listOfTests:
                                myTable.add_row([test["testName"], testType.value, test["url"], test['interval'] , alertsEnabled, test['pageLoadTimeLimit'], test['pageLoadTargetTime'],
                                            test["httpTimeLimit"], test["httpTargetTime"], networkMeasurements, bwMeasurements, bgpMeasurements, mtuMeasurements,
                                            test["protocol"], test["probeMode"], test['pathTraceMode'], test["numPathTraces"], test['sslVersion'],
                                            test["verifyCertificate"], test['authType'], test['httpVersion'], redirects, createdDate, test["enabled"], alertString, agentString])
                                listOfTests.append([test["testName"], testType.value, test["url"], test['interval'] , alertsEnabled, test['pageLoadTimeLimit'], test['pageLoadTargetTime'],
                                            test["httpTimeLimit"], test["httpTargetTime"], networkMeasurements, bwMeasurements, bgpMeasurements, mtuMeasurements,
                                            test["protocol"], test["probeMode"], test['pathTraceMode'], test["numPathTraces"], test['sslVersion'],
                                            test["verifyCertificate"], test['authType'], test['httpVersion'], redirects, createdDate, test["enabled"], alertString, agentString])
                    elif test["type"] == testType.value and test["type"] == TestType.transaction_test.value:
                            if(test["enabled"] == 1):
                                enabled = "Enabled"
                            else:
                                enabled = "Disabled"
                            if test['alertsEnabled'] == 0:
                                alertsEnabled = "No"
                            elif test['alertsEnabled'] == 1:
                                alertsEnabled = "Yes"
                            if test['networkMeasurements'] == 0:
                                networkMeasurements = "No"
                            elif test['networkMeasurements'] == 1:
                                networkMeasurements = "Yes"
                            if test['mtuMeasurements'] == 0:
                                mtuMeasurements = "No"
                            elif test['mtuMeasurements'] == 1:
                                mtuMeasurements = "Yes"
                            if test['bandwidthMeasurements'] == 0:
                                bwMeasurements = "No"
                            elif test['bandwidthMeasurements'] == 1:
                                bwMeasurements = "Yes"
                            if test['bgpMeasurements'] == 0:
                                bgpMeasurements = "No"
                            elif test['bgpMeasurements'] == 1:
                                bgpMeasurements = "Yes"
                            if test['followRedirects'] == 0:
                                redirects = "No"
                            elif test['followRedirects'] == 1:
                                redirects = "Yes"
                            if "createdDate" in test:
                                createdDate = test["createdDate"]
                            
                            alertString = ""
                            agentString = ""
                            #alerts = GetAlertsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                            agents = GetAgentsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                            
                            #for alert in alerts:
                            #    alertString += alert + ","
                            #alertString = alertString[:-1]
                            for agent in agents:
                                agentString += agent + ","
                            agentString = agentString[:-1]
                            if ["web-transaction", test["testName"], test["url"],
                                agentString, test["interval"], "","", "Completed", enabled ] not in results:
                                results.append(["web-transaction", test["testName"], test["url"],
                                                agentString, test["interval"], "","", "Completed", enabled ])    
                            
                    elif test["type"] == testType.value and test["type"] == TestType.voiceRTPStream.value:
                            if(test["enabled"] == 1):
                                enabled = "Enabled"
                            else:
                                enabled = "Disabled"
                            if('server' in test):
                                target = str(test['server']).split(":")[0]
                                port = str(test['server']).split(":")[1]
                            if test['alertsEnabled'] == 0:
                                alertsEnabled = "No"
                            elif test['alertsEnabled'] == 1:
                                alertsEnabled = "Yes"
                            if test['bgpMeasurements'] == 0:
                                bgpMeasurements = "No"
                            elif test['bgpMeasurements'] == 1:
                                bgpMeasurements = "Yes"
                            if "createdDate" in test:
                                createdDate = test["createdDate"]
                            
                            alertString = ""
                            agentString = ""
                            #alerts = GetAlertsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                            agents = GetAgentsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                            
                            #for alert in alerts:
                            #    alertString += alert + ","
                            #alertString = alertString[:-1]
                            for agent in agents:
                                agentString += agent + ","
                            agentString = agentString[:-1]

                            if ["voice (RTP Stream)", test["testName"], target,
                                agentString, test["interval"], "","", "Completed", enabled ] not in results:
                                results.append(["voice (RTP Stream)", test["testName"], target,
                                                agentString, test["interval"], "","", "Completed", enabled ])
                                
                    elif test["type"] == testType.value and test["type"] == TestType.sip_server.value:
                        if(test["enabled"] == 1):
                            enabled = "Enabled"
                        else:
                            enabled = "Disabled"
                        if('server' in test):
                            target = str(test['server']).split(":")[0]
                            port = str(test['server']).split(":")[1]
                        if test['alertsEnabled'] == 0:
                            alertsEnabled = "No"
                        elif test['alertsEnabled'] == 1:
                            alertsEnabled = "Yes"
                        if test['networkMeasurements'] == 0:
                            networkMeasurements = "No"
                        elif test['networkMeasurements'] == 1:
                            networkMeasurements = "Yes"
                        if test['mtuMeasurements'] == 0:
                            mtuMeasurements = "No"
                        elif test['mtuMeasurements'] == 1:
                            mtuMeasurements = "Yes"
                        if test['bgpMeasurements'] == 0:
                            bgpMeasurements = "No"
                        elif test['bgpMeasurements'] == 1:
                            bgpMeasurements = "Yes"
                        if "createdDate" in test:
                                createdDate = test["createdDate"]

                        alertString = ""
                        agentString = ""
                        #alerts = GetAlertsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                        agents = GetAgentsFromTest(myToken=myToken, aid=aid, testID=test["testId"])
                        
                        #for alert in alerts:
                        #    alertString += alert + ","
                        #alertString = alertString[:-1]
                        for agent in agents:
                            agentString += agent + ","
                        agentString = agentString[:-1]

                        if ["sip-server", test["testName"], target,
                                agentString, test["interval"], "","", "Completed", enabled ] not in results:
                                results.append(["sip-server", test["testName"], target,
                                                agentString, test["interval"], "","", "Completed", enabled ])
    else:
        return
    return sorted(results, key=lambda x:x[0])

def GetTestsSummary(myToken:str, testType:TestType, aid:str, returnType="table", selectColumns="ALL"):
    global headers
    global rate_limit
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/tests.json"

    if rate_limit > rate_limit_threshold or rate_limit == -1:
        myTable = PrettyTable()
        result = session.get(url= BASE_URL + get_resource + "?aid=" + aid, headers = headers)
        if testType == TestType.dns_trace:
            myTable.field_names = ["Test Name", "Test Type", "Test ID", "Domain", "Record Type","Transport", "Interval [s]", "Enabled" ]
        elif testType == TestType.dns_dnssec:
            myTable.field_names = ["Test Name", "Test Type", "Test ID", "Domain", "Record Type", "Interval [s]", "Enabled"]
        elif testType == TestType.agent_to_server:
            myTable.field_names = ["Test Name","Test Type", "Test ID", "Target", "Port", "Protocol", "Probe Mode", "Path Trace Mode", "Interval [s]", "Enabled"]
        elif testType == TestType.http_server:
            myTable.field_names = ["Test Name", "Test Type", "Test ID", "URL", "Protocol", "Probe Mode", "Path Trace Mode", "Interval [s]", "Enabled"]
        elif testType == TestType.page_load:
            myTable.field_names = ["Test Name", "Test Type", "Test ID", "URL", "Protocol", "Probe Mode", "Path Trace Mode", "Interval [s]", "Enabled"]
        
        try:
            rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
        except Exception as ex:
            print(ex)
        
        keyword=""

        for test in result.json()["test"]:
            domain = ""
            recordType = ""
            enabled = ""
            target = ""
            port = ""

            if test["type"] == testType.value and test["type"] == TestType.dns_trace.value:
                if('enabled' in test):
                    if test['enabled'] == 0:
                        enabled = "No"
                    elif test['enabled'] == 1:
                        enabled = "Yes"
                if('domain' in test):
                    domain = str(test['domain']).split(" ")[0]
                    recordType = str(test['domain']).split(" ")[1]
                    
                myTable.add_row([test["testName"], testType.value, test['testId'],domain, recordType ,test['dnsTransportProtocol'],
                                 test["interval"], enabled])
            elif test["type"] == testType.value and test["type"] == TestType.dns_dnssec.value:
                if('domain' in test):
                    domain = str(test['domain']).split(" ")[0]
                    recordType = str(test['domain']).split(" ")[1]
                if('enabled' in test):
                    if test['enabled'] == 0:
                        enabled = "No"
                    elif test['enabled'] == 1:
                        enabled = "Yes"

                myTable.add_row([test["testName"], testType.value, test['testId'],domain, recordType, test["interval"],enabled ])
            elif test["type"] == testType.value and test["type"] == TestType.agent_to_server.value:
                if('server' in test):
                    target = str(test['server']).split(":")[0]
                    port = str(test['server']).split(":")[1]
                if('enabled' in test):
                    if test['enabled'] == 0:
                        enabled = "No"
                    elif test['enabled'] == 1:
                        enabled = "Yes"

                myTable.add_row([test["testName"], testType.value, test['testId'],target, port, test['protocol'] , test["probeMode"],
                                test["pathTraceMode"], test["interval"], enabled])
            elif test["type"] == testType.value and test["type"] == TestType.http_server.value:
                if('enabled' in test):
                    if test['enabled'] == 0:
                        enabled = "No"
                    elif test['enabled'] == 1:
                        enabled = "Yes"
                
                myTable.add_row([test["testName"], testType.value, test['testId'],test["url"], test["protocol"], test["probeMode"], test['pathTraceMode'],
                                 test["interval"], enabled])
            elif test["type"] == testType.value and test["type"] == TestType.page_load.value:
                if('enabled' in test):
                    if test['enabled'] == 0:
                        enabled = "No"
                    elif test['enabled'] == 1:
                        enabled = "Yes"
                
                myTable.add_row([test["testName"], testType.value,test['testId'],test["url"], test["protocol"], test["probeMode"], test['pathTraceMode'],
                                 test["interval"], enabled])
    else:
        return
    if "Test Name" in myTable.field_names:
        myTable.sortby = "Test Name"
    myTable.align = "l"
    
    if returnType == "csv":
        return myTable.get_csv_string()
    if selectColumns == "ALL":
        return myTable
    else:
        return SelectColumnsFromPrettyTable(myTable, columns=selectColumns)

def GetTestID(myToken:str, aid:str, returnType="table", column="ALL", selectColumns="ALL"):
    global headers
    global rate_limit
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/tests.json"
    myTable = PrettyTable()

    if rate_limit > rate_limit_threshold or rate_limit == -1:
        result = session.get(url= BASE_URL + get_resource + "?aid=" + aid, headers = headers)
        myTable.field_names = ["Test Name", "Test ID", "Test Type", "Enabled" ]

        try:
            print(bcolors.OKCYAN + "Performed Request, remaining: " + str(rate_limit)+ " requests" + bcolors.ENDC)
            rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
        except Exception as ex:
            print(ex)

        if result.status_code > 199 and result.status_code < 400:
            for test in result.json()["test"]:
                myTable.add_row([test["testName"], test['testId'],test['type'], test["enabled"]])
        
        myTable.sortby = "Test Type"
    
    else:
        return
    #if(column != "ALL"):
    #    columns = column.split(",")
    #    columns = [s.strip() for s in columns]
    #    fieldNames = [str(s).strip() for s in myTable.field_names]
    #    for tableColumn in fieldNames:
    #        if tableColumn not in columns:
    #            myTable.del_column(tableColumn)
    #    if myTable.sortby not in myTable.field_names:
    #        myTable.sortby = myTable.field_names[0]
    myTable.align = "l"

    if returnType == "csv":
        return myTable.get_csv_string()
    if selectColumns == "ALL":
        return myTable
    else:
        return SelectColumnsFromPrettyTable(myTable, columns=selectColumns)   

def DeleteTestID(myToken:str, aid:str, id):
    global headers
    global rate_limit
    listOfSuccessIDs = []
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    myTable = PrettyTable()

    if rate_limit > rate_limit_threshold or rate_limit == -1:
        get_resource = "/tests/" + str(id) + ".json"

        result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)

        try:
            rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
        except Exception as ex:
            print(ex)

        if result.status_code > 199 and result.status_code < 400:
            jsonResult = json.loads(result.text)['test']
            testType = jsonResult[0]['type']
            post_resource = "/tests/" + str(testType) + "/" + str(id) + "/delete.json"
            try:
                result = session.post(url= BASE_URL + post_resource + "?aid=" + str(aid), headers = headers)
                if result.status_code > 199 and result.status_code < 400:
                    print(bcolors.OKGREEN + bcolors.BOLD+ str(id) + " : SUCCESS : " + "deleted from account ID " + str(aid) + bcolors.ENDC)
                    listOfSuccessIDs.append(id)
                else:
                    print(bcolors.FAIL + str(id) + " : FAIL : " + json.loads(result.text)['errorMessage'] + bcolors.ENDC)
            except Exception as ex:
                print(bcolors.FAIL + str(id) + " : FAIL : " + str(ex.args[0]) + bcolors.ENDC)
        elif result.status_code > 399:
            print(bcolors.FAIL + str(id) + " : FAIL : " + json.loads(result.text)['errorMessage'] + bcolors.ENDC)
        
    else:
        return
    return listOfSuccessIDs

def SetRateLimiteThreshold(threshold):
    global rate_limit_threshold

    if type(threshold) == type(""):
        try:
            rate_limit_threshold = int(threshold)
            return True
        except:
            return False
    if threshold > 1 and threshold < organization_max_rate:
        rate_limit_threshold = threshold
        return True
    return False

def GetActiveAlerts(myToken:str, aid:str, myParameters=[""], returnType = "table", selectColumns = "ALL"):
    """

    """
    global headers
    global rate_limit
    start = ""
    end = ""
    window = ""
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/alerts.json"
    myTable = PrettyTable()

    if len(myParameters) == 1:
        if str(myParameters[0]).strip().endswith("d") or (myParameters[0]).strip().endswith("h") or (myParameters[0]).strip().endswith("m"):
            if str(myParameters[0]).strip().replace("d", "").replace("h","").replace("m","").isnumeric():
                window = myParameters[0]
                end = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
                start = (datetime.today() - timedelta(hours=int(myParameters[0].strip().replace("h", "")))).strftime("%Y-%m-%dT%H:%M:%S")
    elif len(myParameters) == 2:
        start = myParameters[0]
        end = myParameters[1]
    
    if rate_limit > rate_limit_threshold or rate_limit == -1:
        requestString = ""
        if window != "":
            requestString = BASE_URL + get_resource + "?aid=" + str(aid) + "&window=" + window
            result = session.get(url= requestString, headers = headers)
        elif start != "" and end != "":
            requestString = BASE_URL + get_resource + "?aid=" + str(aid) + "&from=" + start + "&to=" + end
            result = session.get(url=requestString , headers = headers)
        
        myTable.field_names = ["Alert Name", "Alert ID", "Alert Type", "Violation Count", "Rule", "Tests" ]

        if result.status_code > 199 and result.status_code < 399:
            for alert in result.json()["alert"]:
                myTable.add_row([alert['ruleName'], alert['alertId'], alert['type'], alert['violationCount'], alert['ruleExpression'], alert['testName']])
    myTable.align = "l"
    if returnType == "csv":
        return myTable.get_csv_string()
    if selectColumns == "ALL":
        return myTable
    else:
        return SelectColumnsFromPrettyTable(myTable, columns=selectColumns)
    
    return myTable

def GetListOfUsers(myToken:str, aid:str, returnType="table", selectColumns="ALL"):
    global headers
    global rate_limit
    myTable = PrettyTable()
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/users.json"

    if rate_limit > rate_limit_threshold or rate_limit == -1:
        myTable.field_names = ["User Name", "Email", "Last Login", "Registered on"]
        result = session.get(url= BASE_URL + get_resource + "?aid=" + aid, headers = headers)
        
        try:
            rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
        except Exception as ex:
            print(ex)

        for user in result.json()['users']:
            myTable.add_row([user["name"],user["email"],user["lastLogin"],user["dateRegistered"]])
            
        myTable.sortby = "User Name"
        myTable.align = "l"
    
        if returnType == "csv":
            return myTable.get_csv_string()
        if selectColumns == "ALL":
            return myTable
        else:
            return SelectColumnsFromPrettyTable(myTable, columns=selectColumns)
    code = result.status_code 

def DisableTest(myToken:str, aid:str, id):
    global headers
    global rate_limit
    listOfSuccessIDs = []
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    myTable = PrettyTable()

    if rate_limit > rate_limit_threshold or rate_limit == -1:
        get_resource = "/tests/" + str(id) + ".json"

        result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)

        try:
            rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
        except Exception as ex:
            print(ex)

        if result.status_code > 199 and result.status_code < 400:
            jsonResult = json.loads(result.text)['test']
            testType = jsonResult[0]['type']
            post_resource = "/tests/" + str(testType) + "/" + str(id) + "/update.json"
            try:
                payload = { "enabled" : 0 }
                result = session.post(url= BASE_URL + post_resource + "?aid=" + str(aid), data = json.load(payload), headers = headers)
                if result.status_code > 199 and result.status_code < 400:
                    print(bcolors.OKGREEN + bcolors.BOLD+ str(id) + " : SUCCESS : " + "disabled test " + str(jsonResult[0]['testName']) + bcolors.ENDC)
                    listOfSuccessIDs.append(id)
                else:
                    print(bcolors.FAIL + str(id) + " : FAIL : " + str(jsonResult[0]['testName']) + json.loads(result.text)['errorMessage'] + bcolors.ENDC)
            except Exception as ex:
                print(bcolors.FAIL + str(id) + " : FAIL : " + str(ex.args[0]) + bcolors.ENDC)
        elif result.status_code > 399:
            print(bcolors.FAIL + str(id) + " : FAIL : " + json.loads(result.text)['errorMessage'] + bcolors.ENDC)
        
    else:
        return
    return listOfSuccessIDs

def GetQuotaUtilization(myToken:str, aid:str):
    global headers
    global rate_limit
    myTable = PrettyTable()
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/usage.json"

    result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)
    try:
        rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
    except Exception as ex:
        print(ex)

    if result.status_code > 199 and result.status_code < 400:
        jsonResult = json.loads(result.text)['usage']
        quotaUnits = jsonResult["quota"]["cloudUnitsIncluded"]
        unitsUsed = int(jsonResult["cloudUnitsNextBillingPeriod"]) + int(jsonResult["enterpriseUnitsNextBillingPeriod"])

        return unitsUsed/quotaUnits

    else:
        return -1

def GetOrganizationName(myToken:str, aid:str):
    global headers
    global rate_limit
    myTable = PrettyTable()
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/account-groups/" + str(aid) + ".json"

    result = session.get(url= BASE_URL + get_resource, headers = headers)
    try:
        rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
    except Exception as ex:
        print(ex)

    if result.status_code > 199 and result.status_code < 400:
        jsonResult = json.loads(result.text)['accountGroups']
        for accountGroup in jsonResult:
            if accountGroup["aid"] == aid:
                return accountGroup["organizationName"]

def GetDashboardDetail(myToken:str, aid:str, dashboardId:str):
    global headers
    global rate_limit
    myTable = PrettyTable()
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/dashboards/" + dashboardId + ".json"
    result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)
    if(result.status_code == 429):
        print(bcolors.WARNING + "Reached Rate limit: Time to refresh " + str((datetime.fromtimestamp(int(result.headers["x-organization-rate-limit-reset"])) - datetime.now()).seconds) + " seconds" + bcolors.ENDC)
        while(result.status_code == 429):
            result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)
    try:
        rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
        print(bcolors.OKCYAN + "Performed Request, remaining: " + str(rate_limit)+ " requests" + bcolors.ENDC)
    except Exception as ex:
        print(ex)

    if result.status_code > 199 and result.status_code < 400:
        return json.loads(result.text)["dashboards"]

def GetDashboards(myToken:str, aid:str, reportType:DocumentType, removeBuiltIn=True):
    global headers
    global rate_limit
    myTable = PrettyTable()
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/dashboards.json"
    dashboardsEnterprise = []
    dashboardsEndpoint = []

    result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)
    if(result.status_code == 429):
        print(bcolors.WARNING + "Reached Rate limit: Time to refresh " + str((datetime.fromtimestamp(int(result.headers["x-organization-rate-limit-reset"])) - datetime.now()).seconds) + " seconds" + bcolors.ENDC)
        while(result.status_code == 429):
            result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)
    try:
        rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
        print(bcolors.OKCYAN + "Performed Request, remaining: " + str(rate_limit)+ " requests" + bcolors.ENDC)
    except Exception as ex:
        print(ex)

    if result.status_code > 199 and result.status_code < 400:
        jsonResult = json.loads(result.text)["dashboards"]
        for dashboard in jsonResult:
            hasEndpointData = False
            hasEnterpriseData = False
            dashboardDetail = GetDashboardDetail(myToken, aid, dashboard["id"])
            for widget in dashboardDetail[0]["widgets"]:
                for component in widget["dataComponents"]:
                    if "metric" in component:
                        if "Endpoint" in component["metric"]:
                            hasEndpointData = True
                        else:
                            hasEnterpriseData = True
            
            if removeBuiltIn:
                if hasEndpointData and not dashboard["isBuiltIn"]:
                    #Dashboard Name     Number of Widgets   
                    dashboardsEndpoint.append([dashboardDetail[0]["name"], str(len(dashboardDetail[0]["widgets"]))])
                if hasEnterpriseData and not dashboard["isBuiltIn"]:
                    #Dashboard Name     Number of Widgets   
                    dashboardsEnterprise.append([dashboardDetail[0]["name"], str(len(dashboardDetail[0]["widgets"]))])
            else:
                if hasEndpointData:
                    #Dashboard Name     Number of Widgets   IsBuiltIn
                    dashboardsEndpoint.append([dashboardDetail[0]["name"], str(len(dashboardDetail[0]["widgets"])), dashboard["isBuiltIn"]])
                if hasEnterpriseData:
                    #Dashboard Name     Number of Widgets   IsBuiltIn
                    dashboardsEnterprise.append([dashboardDetail[0]["name"], str(len(dashboardDetail[0]["widgets"])), dashboard["isBuiltIn"]])
        
        if reportType == DocumentType.ENDPOINT_AGENTS:
            return dashboardsEndpoint
        elif reportType == DocumentType.ENTERPRISE_AGENTS:
            return dashboardsEnterprise

def GetAlertDetails(myToken:str, aid:str, alertId):
    global headers
    global rate_limit
    myTable = PrettyTable()
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/alert-rules/" + str(alertId) + ".json"
    dashboardsEnterprise = []
    dashboardsEndpoint = []

    result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)
    if(result.status_code == 429):
        print(bcolors.WARNING + "Reached Rate limit: Time to refresh " + str((datetime.fromtimestamp(int(result.headers["x-organization-rate-limit-reset"])) - datetime.now()).seconds) + " seconds" + bcolors.ENDC)
        while(result.status_code == 429):
            result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)
    try:
        rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
        print(bcolors.OKCYAN + "Performed Request, remaining: " + str(rate_limit)+ " requests" + bcolors.ENDC)
    except Exception as ex:
        print(ex)

    if result.status_code > 199 and result.status_code < 400:
        jsonResult = json.loads(result.text)["alertRules"]
        return jsonResult[0]
    else:
        return []

def GetListOfAlerts(myToken:str, aid:str, reportType:DocumentType):
    global headers
    global rate_limit
    myTable = PrettyTable()
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/alert-rules.json"
    
    listOfAlerts = []

    result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)
    if(result.status_code == 429):
        print(bcolors.WARNING + "Reached Rate limit: Time to refresh " + str((datetime.fromtimestamp(int(result.headers["x-organization-rate-limit-reset"])) - datetime.now()).seconds) + " seconds" + bcolors.ENDC)
        while(result.status_code == 429):
            result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)
    try:
        rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
        print(bcolors.OKCYAN + "Performed Request, remaining: " + str(rate_limit)+ " requests" + bcolors.ENDC)
    except Exception as ex:
        print(ex)

    if result.status_code > 199 and result.status_code < 400:
        jsonResult = json.loads(result.text)["alertRules"]
    
    for alertRule in jsonResult:
        expression = str(re.sub(r"(.)(?=.*\1)\)", ")", alertRule["expression"]))
        expression = expression.replace("||", "\nor\n").replace("&&", "and").replace("!", " not ").replace("(","").replace(")", ","). replace("not = \"None\"", "not present")
        expression = str(re.sub(r"(.)(?=.*\1)\)", ")", expression))
        severity = GetAlertDetails(myToken, aid, alertRule["ruleId"])["severity"]
        if "minimumSourcesPct" in alertRule:
            percentageLegend = str(alertRule["minimumSourcesPct"]) + "% "
        else:
            percentageLegend = str(alertRule["minimumSources"]) + " "

        if "roundsViolatingMode" in alertRule:
            listOfAlerts.append([alertRule["alertType"],alertRule["ruleName"], "", expression, str(alertRule["roundsViolatingMode"]) + " of " + percentageLegend + 
                                                                                                "agent(s), and " + str(alertRule["roundsViolatingRequired"]) + 
                                                                                                " out of " + str(alertRule["roundsViolatingOutOf"]) + " times in a row", severity, ""])
        else:
            listOfAlerts.append([alertRule["alertType"],alertRule["ruleName"], "", expression, "Conditions met by at least " + percentageLegend +
                                                                                                "agent(s), and " + str(alertRule["roundsViolatingRequired"]) + 
                                                                                                " out of " + str(alertRule["roundsViolatingOutOf"]) + " times in a row", severity, ""])
    
    uniqueValues = []

    for alert in listOfAlerts:
        if alert not in uniqueValues:
            uniqueValues.append(alert)
    
    return uniqueValues

def GetLabelsFromIds(myToken:str, listOfLabelIds, aid:str):
    global headers
    global rate_limit
    myTable = PrettyTable()
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/groups.json"
    results = []

    if rate_limit > rate_limit_threshold or rate_limit == -1:
        result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)
        if(result.status_code == 429):
            print(bcolors.WARNING + "Reached Rate limit: Time to refresh " + str((datetime.fromtimestamp(int(result.headers["x-organization-rate-limit-reset"])) - datetime.now()).seconds) + " seconds" + bcolors.ENDC)
            while(result.status_code == 429):
                result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)

        try:
            rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
            print(bcolors.OKCYAN + "Performed Request, remaining: " + str(rate_limit)+ " requests" + bcolors.ENDC)
        except Exception as ex:
            print(ex)
    for label in result.json()["groups"]:
        if label["groupId"] in listOfLabelIds:
            results.append([label["groupId"], label["name"]])
    return results

def GetAgentsFromIds(myToken:str, listOfAgentIds, aid:str):
    global headers
    global rate_limit
    myTable = PrettyTable()
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/endpoint-agents.json"
    results = []

    if rate_limit > rate_limit_threshold or rate_limit == -1:
        result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)
        if(result.status_code == 429):
            print(bcolors.WARNING + "Reached Rate limit: Time to refresh " + str((datetime.fromtimestamp(int(result.headers["x-organization-rate-limit-reset"])) - datetime.now()).seconds) + " seconds" + bcolors.ENDC)
            while(result.status_code == 429):
                result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)

        try:
            rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
            print(bcolors.OKCYAN + "Performed Request, remaining: " + str(rate_limit)+ " requests" + bcolors.ENDC)
        except Exception as ex:
            print(ex)
    for agent in result.json()["endpointAgents"]:
        if agent["agentId"] in listOfAgentIds:
            results.append([agent["agentId"], agent["agentName"]])
    return results


def GetEndpointTests(myToken:str, testType:TestType, aid:str):
    global headers
    global rate_limit
    myTable = PrettyTable()
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/" + testType.value + ".json"
    results = []

    if rate_limit > rate_limit_threshold or rate_limit == -1:
        result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)
        if(result.status_code == 429):
            print(bcolors.WARNING + "Reached Rate limit: Time to refresh " + str((datetime.fromtimestamp(int(result.headers["x-organization-rate-limit-reset"])) - datetime.now()).seconds) + " seconds" + bcolors.ENDC)
            while(result.status_code == 429):
                result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)

        try:
            rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
            print(bcolors.OKCYAN + "Performed Request, remaining: " + str(rate_limit)+ " requests" + bcolors.ENDC)
        except Exception as ex:
            print(ex)
        
        if(testType == TestType.scheduled_test):
            assignedTo = ""
            status = ""
            for test in result.json()["endpointTest"]:
                if test["agentSelectorConfig"]["agentSelectorType"] == "LIST_OF_LABELS":
                    listOfLabels = GetLabelsFromIds(myToken, test["agentSelectorConfig"]["labelIds"], aid)
                    for label in listOfLabels:
                        assignedTo += str(label[1]) + ","
                    
                    assignedTo = assignedTo[:-1]

                elif test["agentSelectorConfig"]["agentSelectorType"] == "ANY_AGENT":
                    assignedTo = "All Endpoint Agents"
                elif test["agentSelectorConfig"]["agentSelectorType"] == "SPECIFIC_AGENTS":
                    listOfAgents = GetAgentsFromIds(myToken, test["agentSelectorConfig"]["agentIds"], aid)
                    #for agent in listOfAgents:
                        #assignedTo += str(label[1]) + ","
                    
                    #assignedTo = assignedTo[:-1]
                    assignedTo = "SPECIFIC_AGENTS" + " (Total: " + str(len(listOfAgents)) + ")"
                
                if test["enabled"] == 1:
                    status = "Enabled"
                if test["enabled"] == 0:
                    status = "Disabled"

                if "savedEvent" in test:
                    if test["savedEvent"] == 0:
                        results.append(["Scheduled - " + str(test["type"]), test["testName"], test["server"] , test["interval"], assignedTo, "", status])
        elif(testType == TestType.automated_test):
            assignedTo = ""
            status = ""
            for test in result.json()["automatedSessionTests"]:
                if test["agentSelectorConfig"]["agentSelectorType"] == "LIST_OF_LABELS":
                    listOfLabels = GetLabelsFromIds(myToken, test["agentSelectorConfig"]["labelIds"], aid)
                    for label in listOfLabels:
                        assignedTo += str(label[1]) + ","
                    
                    assignedTo = assignedTo[:-1]

                elif test["agentSelectorConfig"]["agentSelectorType"] == "ANY_AGENT":
                    assignedTo = "All Endpoint Agents"
                elif test["agentSelectorConfig"]["agentSelectorType"] == "SPECIFIC_AGENTS":
                    listOfAgents = GetAgentsFromIds(myToken, test["agentSelectorConfig"]["agentIds"], aid)
                    #for agent in listOfAgents:
                    #    assignedTo += str(label[1]) + ","
                    
                    #assignedTo = assignedTo[:-1]
                    assignedTo = "SPECIFIC_AGENTS" + " (Total: " + str(len(listOfAgents)) + ")"
                
                if test["enabled"] == 1:
                    status = "Enabled"
                if test["enabled"] == 0:
                    status = "Disabled"

                if "savedEvent" in test:
                    if test["savedEvent"] == 0:
                        if "type" in test:
                            results.append(["Dynamic - " + str(test["type"]), test["testName"], test["application"], test["interval"], assignedTo, "", status])
                        else:
                            results.append(["Dynamic - " + str(test["protocol"]), test["testName"], test["application"] , test["interval"], assignedTo, "", status])
    
    return results

def GetEndpointLicensing(myToken:str, aid:str, endpointLicense:EndpointLicense):
    global headers
    global rate_limit
    myTable = PrettyTable()
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/usage.json"

    result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)
    if(result.status_code == 429):
        print(bcolors.WARNING + "Reached Rate limit: Time to refresh " + str((datetime.fromtimestamp(int(result.headers["x-organization-rate-limit-reset"])) - datetime.now()).seconds) + " seconds" + bcolors.ENDC)
        while(result.status_code == 429):
            result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)
    try:
        rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
        print(bcolors.OKCYAN + "Performed Request, remaining: " + str(rate_limit)+ " requests" + bcolors.ENDC)
    except Exception as ex:
        print(ex)

    if result.status_code > 199 and result.status_code < 400:
        jsonResult = json.loads(result.text)["usage"]["quota"]
        if endpointLicense==EndpointLicense.ADVANTAGE:
            if "endpointAgentsIncluded" in jsonResult:
                return jsonResult["endpointAgentsIncluded"]
            else:
                return "0"
        elif endpointLicense == EndpointLicense.ESSENTIALS:
            if "endpointAgentsEssentialsIncluded" in jsonResult:
                return jsonResult["endpointAgentsEssentialsIncluded"]
            else:
                return "0"
        elif endpointLicense == EndpointLicense.EMBEDDED:
            if "endpointAgentsEmbeddedIncluded" in jsonResult:
                return jsonResult["endpointAgentsEmbeddedIncluded"]
            else:
                return "0"

def GetEndpointLicenseUsage(myToken:str, aid:str, endpointLicense:EndpointLicense):
    global headers
    global rate_limit
    myTable = PrettyTable()
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer ' + myToken,
    }
    get_resource = "/usage.json"

    result = session.get(url= BASE_URL + get_resource + "?aid=" + str(aid), headers = headers)
    if(result.status_code == 429):
        print(bcolors.WARNING + "Reached Rate limit: Time to refresh " + str((datetime.fromtimestamp(int(result.headers["x-organization-rate-limit-reset"])) - datetime.now()).seconds) + " seconds" + bcolors.ENDC)
        while(result.status_code == 429):
            result = session.get(url= BASE_URL + get_resource, headers = headers)
    try:
        rate_limit = int(result.headers["x-organization-rate-limit-remaining"])
        print(bcolors.OKCYAN + "Performed Request, remaining: " + str(rate_limit)+ " requests" + bcolors.ENDC)
    except Exception as ex:
        print(ex)

    if result.status_code > 199 and result.status_code < 400:
        jsonResult = json.loads(result.text)['usage']
        if endpointLicense==EndpointLicense.ADVANTAGE:
            return jsonResult["endpointAgents"][0]["endpointAgentsUsed"]
        elif endpointLicense == EndpointLicense.ESSENTIALS:
            return jsonResult["endpointAgentsEssentials"][0]["endpointAgentsEssentialsUsed"]
        elif endpointLicense == EndpointLicense.EMBEDDED:
            return jsonResult["endpointAgentsEmbedded"][0]["endpointAgentsEmbeddedUsed"]

def BuildImplementationPlan(templatePath:str, myToken:str, aid:str):
    #Read implementation file
    shutil.copyfile(templatePath, "tempImplementationPlan.xlsx")
    fileToUse = "tempImplementationPlan.xlsx"

    workbook = load_workbook(filename=fileToUse)
    summarySheet = workbook["Summary"]
    testsSheet = workbook["Tests Details"]
    agentSheet = workbook["Agent Site List"]
    endpointSheet = workbook["Endpoint Details"]
    alertsSheet = workbook["Alerts & Dashboards"]

    #region "Summary tab"
    print(bcolors.OKGREEN + "Working on: Summary tab")
    orgName = GetOrganizationName(myToken, aid)
    i=0
    for r in range(1,summarySheet.max_row+1):
        for c in range(1,summarySheet.max_column+1):
            s = str(summarySheet.cell(r,c).value)
            if s != None and "[CLIENT]" in s: 
                summarySheet.cell(r,c).value = s.replace("[CLIENT]",str(orgName)) 
                i += 1
    
    summarySheet["C12"] = orgName
    today = datetime.now()
    summarySheet["C15"] = today.strftime("%B") + " " + str(today.day) + "," + str(today.year)
    summarySheet["B28"] = today.strftime("%B") + " " + str(today.day) + "," + str(today.year)
    summarySheet["C28"] = "Initial Version"
    summarySheet["E28"] = "0.1"
    #endregion

    #region "Agent List"
    print(bcolors.OKGREEN + "Working on: Agents tab")
    clusterAgentRows = GetListOfAgents(myToken, AgentType.ENTERPRISE_CLUSTER, aid)
    enterpriseAgentRows = GetListOfAgents(myToken, AgentType.ENTERPRISE, aid)
    endpointAgentsRows = GetListOfAgents(myToken, AgentType.ENDPOINT, aid)
    agentRow=0
    agentCol=0

    #Write Agent Data into Excel file
    clusterRows_count = len(clusterAgentRows)-1
    if(clusterRows_count >= 0):
        for row in agentSheet.iter_rows(min_row=9, max_row=9+clusterRows_count, min_col=2, max_col=7):
            for value in clusterAgentRows[agentRow]:
                row[agentCol].value = value
                agentCol += 1
            agentCol = 0
            agentRow+=1
    
    agentRow=0
    agentCol = 0

    enterpriseRows_count = len(enterpriseAgentRows)-1
    if(enterpriseRows_count >= 0):
        for row in agentSheet.iter_rows(min_row=9+clusterRows_count+1, max_row=9+clusterRows_count+len(enterpriseAgentRows), min_col=2, max_col=7):
            for value in enterpriseAgentRows[agentRow]:
                row[agentCol].value = value
                agentCol += 1
            agentCol = 0
            agentRow+=1
    
    agentRow=0
    agentCol = 0
    endpointRows_count = len(endpointAgentsRows)-1
    if(endpointRows_count >= 0):
        for row in agentSheet.iter_rows(min_row=9+clusterRows_count+len(enterpriseAgentRows)+1, max_row=9+clusterRows_count+len(enterpriseAgentRows)+len(endpointAgentsRows), min_col=2, max_col=7):
            for value in endpointAgentsRows[agentRow]:
                row[agentCol].value = value
                agentCol += 1
            agentCol = 0
            agentRow+=1

    #endregion
    
    #region "Tests and Unit Consumption"

    print(bcolors.OKGREEN + "Working on: Tests tab")
    unitConsumption = GetQuotaUtilization(myToken, aid)
    testsSheet["C7"] = round(unitConsumption, 4)

    listOfTests = GetTests(myToken, TestType.dns_server, aid)
    listOfTests += (GetTests(myToken, TestType.dns_trace, aid))
    listOfTests += (GetTests(myToken, TestType.dns_dnssec, aid))
    listOfTests += (GetTests(myToken, TestType.agent_to_server, aid))
    listOfTests += (GetTests(myToken, TestType.agent_to_agent, aid))
    listOfTests += (GetTests(myToken, TestType.voiceRTPStream, aid))
    listOfTests += (GetTests(myToken, TestType.sip_server, aid))
    listOfTests += (GetTests(myToken, TestType.http_server, aid))
    listOfTests += (GetTests(myToken, TestType.page_load, aid))
    listOfTests += (GetTests(myToken, TestType.transaction_test, aid))

    #Fill the table
    testRow = 0
    testCol = 0
    Rows_count = 0

    Rows_count = len(listOfTests)-1
    for row in testsSheet.iter_rows(min_row=11, max_row=11+Rows_count, min_col=2, max_col=10):
        if(len(listOfTests[testRow])>0):
            for value in listOfTests[testRow]:
                row[testCol].value = value
                testCol += 1
        testCol = 0
        testRow += 1
        
    #endregion
    
    #region "EPA Licenses, Scheduled Tests and Dynamic Tests"
    
    print(bcolors.OKGREEN + "Working on: Endpoint data tab")
    advantageConsumption = str(GetEndpointLicenseUsage(myToken, aid, EndpointLicense.ADVANTAGE)) + " out of " + str(GetEndpointLicensing(myToken, aid, EndpointLicense.ADVANTAGE))
    essentialsConsumption = str(GetEndpointLicenseUsage(myToken, aid, EndpointLicense.ESSENTIALS)) + " out of " + str(GetEndpointLicensing(myToken, aid, EndpointLicense.ESSENTIALS))
    endpointSheet["C9"] = advantageConsumption
    endpointSheet["C10"] = essentialsConsumption
    listOfTests = GetEndpointTests(myToken, TestType.scheduled_test, aid)
    listOfTests += (GetEndpointTests(myToken, TestType.automated_test, aid))

    #Fill the table
    testRow = 0
    testCol = 0
    Rows_count = 0

    Rows_count = len(listOfTests)-1
    for row in endpointSheet.iter_rows(min_row=24, max_row=24+Rows_count, min_col=2, max_col=8):
        if(len(listOfTests[testRow])>0):
            for value in listOfTests[testRow]:
                row[testCol].value = value
                testCol += 1
        testCol = 0
        testRow += 1

    #endregion
        
    #region "Alerts"
    
    print(bcolors.OKGREEN + "Working on: Alerts")
    listOfAlerts = GetListOfAlerts(myToken, aid, DocumentType.ENTERPRISE_AGENTS)
    listOfAlerts += (GetListOfAlerts(myToken, aid, DocumentType.ENDPOINT_AGENTS))

    #Fill the table
    testRow = 0
    testCol = 0
    Rows_count = 0

    Rows_count = len(listOfAlerts)-1
    for row in alertsSheet.iter_rows(min_row=11, max_row=11+Rows_count, min_col=2, max_col=8):
        if(len(listOfAlerts[testRow])>0):
            for value in listOfAlerts[testRow]:
                row[testCol].value = value
                testCol += 1
        testCol = 0
        testRow += 1

    #endregion

    workbook.save(orgName + "_ImplementationPlan.xlsx")




def main():
    global token
    global CMDs
    global aid
    cmd = ""
    prompt = ">"
    prePend = "(Rate-limit-remaining:"+str(rate_limit)+")"

    print(bcolors.OKCYAN,"""
                                                                                                                                                                          
                                                                                                                                                                          
                        ...                                ..                                                           ....       ..:::::::..                            
                       :===                               ===:                                                    .:==++=:    .-=+*************+=-:                       
                       :===                               ===:                                                 .-+***+:   .:=***********************+-.                   
               ...     :===      ..               ..      ===:     ...                                      .-+***+-.  .-+*********++========+*********+-.                
               ===.    :===     :==:             .===     ===:     ====                                    :+****+:   .=********+-:             .:=+*******=:              
       ..      ===.    :===     :==:      .      .===     ===:     ===.     ..                          :+****+-   :=*******=:        .-+++==-.    :=*******+:            
      ===:     ===.    :===     :==:     ===.    .===     ===:     ===.    :===                       :+****+-   .+******+-.             .-+****+-.   :=******+.          
      ===:     ===.    :===     :==:    .===.    .===     ===:     ===.    :===                      =*****=    =******=.         .          -+****+:   .+******=.        
      :==:     :==.    :===     .==.     ===.    .==:     ===:     ====     .==:                    :+****+-   :+*****+:        :+###*=.        :+****+-   :+******-       
                       :===                               ===:                                    -*****=.   =******-         :#######*          :+****+:   :+*****+.     
                        .:.                                ::                                     :*****=.   -******=.        :#######*          :+****+.   -******=      
                                                                                                   .=****+-   .=******-        :+*##*=.        -+****+:   -+*****+:       
                                                                                                     -*****=.   :+*****+-                   .-*****=.   -+******-         
               .:=====:    ====     .:====:      .:=====:      :=====:.                                .=*****-   .=******+-.              :=****=:.   -+******=.          
             .=====::=:    ===.   :===::::    .=====::=:    :====:=====.                                .=****+-.  .=*******=:.        :====-:.    :=*******=:            
             ===:          ===.   ====:.      ====.        ====     :===                                  .=*****-.   -********+=:.           .:=+********-.              
            .===.          ===.    :======.   ===:         ===:     .===.                                   .-+****=.   :+*********++==---==++*********+-                 
             :===.         ===.       .:===   :===:    .   :===.   .====                                       .-=***+-.  .:=+*********************+=-.                   
              :=======:    ===.   ========:    .========    :=========:                                            .:=+++-.   .:==+************+=-.                       
                .:::::.    :::    ::::::.        ..::::.      .:::::.                                                    ...        ..:::::..                             
                                                                                                                                                                          
                                                                                                                                                                          
                                                                                                                                                                    """, bcolors.ENDC)
    dummyTable = PrettyTable()
    dummyTable.add_row(["Coded by: lusarmie\n"])
    print(bcolors.HEADER + "Coded by: lusarmie", bcolors.ENDC)
    print(bcolors.OKGREEN + bcolors.BOLD,"Welcome to TE Plan Builder !\n", bcolors.ENDC)
    token = input("Please provide your token: ")
    
    while(not TestCredentials(token)):
        print("Could not validate your token, please verify.")
        token = input("Please provide your token: ")
    
    while(ValidateAccountName(token, "") == False):
        print()

    startTime = datetime.now()
    BuildImplementationPlan("Thousandeyes - Implementation Plan_2023.xlsx", token, aid)
    print(bcolors.BOLD + bcolors.OKGREEN, "Time elapsed: " + str((datetime.now() - startTime).seconds) + " seconds")
    watchdog.stop()

if __name__ == "__main__":
    main()
